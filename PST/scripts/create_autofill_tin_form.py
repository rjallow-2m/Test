"""
Create ODK XLSForm with CSV lookup based on TIN
When staff enter their TIN, personal information auto-fills from staff-list-with-gender.csv
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Define the survey structure with TIN lookup
survey_data = [
    # TIN Input Field
    {
        'type': 'text',
        'name': 'tin',
        'label': 'Enter your TIN (Tax Identification Number)',
        'required': 'yes',
        'hint': 'Enter your TIN to auto-fill your information'
    },
    
    # Auto-filled fields from CSV using pulldata()
    # Personal Information Group
    {'type': 'begin_group', 'name': 'personal_info', 'label': 'Personal Information'},
    {
        'type': 'calculate',
        'name': 'eeno',
        'label': 'Employee Number',
        'calculation': 'pulldata(\'staff_list\', \'eeno\', \'tin\', ${tin})'
    },
    {
        'type': 'calculate',
        'name': 'full_name',
        'label': 'Full Name',
        'calculation': 'pulldata(\'staff_list\', \'staff_name\', \'tin\', ${tin})'
    },
    {
        'type': 'calculate',
        'name': 'gender',
        'label': 'Gender',
        'calculation': 'pulldata(\'staff_list\', \'gender\', \'tin\', ${tin})'
    },
    {
        'type': 'calculate',
        'name': 'organisation',
        'label': 'Organisation',
        'calculation': 'pulldata(\'staff_list\', \'organisation\', \'tin\', ${tin})'
    },
    {
        'type': 'calculate',
        'name': 'job',
        'label': 'Job Title/Position',
        'calculation': 'pulldata(\'staff_list\', \'job\', \'tin\', ${tin})'
    },
    {'type': 'end_group'},
    
    # Display note to show if TIN was found
    {
        'type': 'note',
        'name': 'verification_note',
        'label': 'Staff Found: ${full_name}',
        'relevant': '${full_name} != ""'
    },
    {
        'type': 'note',
        'name': 'not_found_note',
        'label': 'TIN not found in system. Please verify your TIN.',
        'relevant': '${full_name} = ""'
    },
    
    # Additional data collection fields
    {'type': 'begin_group', 'name': 'contact_info', 'label': 'Contact Information', 'relevant': '${full_name} != ""'},
    {'type': 'text', 'name': 'phone_number', 'label': 'Phone Number', 'required': 'yes'},
    {'type': 'text', 'name': 'email', 'label': 'Email Address', 'required': 'yes'},
    {'type': 'text', 'name': 'residential_address', 'label': 'Residential Address', 'required': 'yes'},
    {'type': 'text', 'name': 'city', 'label': 'City', 'required': 'yes'},
    {'type': 'text', 'name': 'postal_code', 'label': 'Postal Code', 'required': 'no'},
    {'type': 'end_group'},
    
    {'type': 'begin_group', 'name': 'emergency_contact', 'label': 'Emergency Contact', 'relevant': '${full_name} != ""'},
    {'type': 'text', 'name': 'emergency_contact_name', 'label': 'Emergency Contact Name', 'required': 'yes'},
    {'type': 'text', 'name': 'emergency_contact_relationship', 'label': 'Relationship', 'required': 'yes'},
    {'type': 'text', 'name': 'emergency_contact_phone', 'label': 'Emergency Contact Phone', 'required': 'yes'},
    {'type': 'end_group'},
    
    {'type': 'begin_group', 'name': 'education_skills', 'label': 'Education & Skills', 'relevant': '${full_name} != ""'},
    {'type': 'select_one education_level', 'name': 'highest_education', 'label': 'Highest Level of Education', 'required': 'yes'},
    {'type': 'text', 'name': 'field_of_study', 'label': 'Field of Study', 'required': 'yes'},
    {'type': 'text', 'name': 'certifications', 'label': 'Professional Certifications', 'required': 'no'},
    {'type': 'text', 'name': 'key_skills', 'label': 'Key Skills', 'required': 'yes'},
    {'type': 'end_group'},
    
    {'type': 'begin_group', 'name': 'additional_info', 'label': 'Additional Information', 'relevant': '${full_name} != ""'},
    {'type': 'date', 'name': 'date_of_birth', 'label': 'Date of Birth', 'required': 'yes'},
    {'type': 'text', 'name': 'nationality', 'label': 'Nationality', 'required': 'yes'},
    {'type': 'select_one marital_status', 'name': 'marital_status', 'label': 'Marital Status', 'required': 'yes'},
    {'type': 'date', 'name': 'start_date', 'label': 'Employment Start Date', 'required': 'yes'},
    {'type': 'integer', 'name': 'years_of_experience', 'label': 'Total Years of Work Experience', 'required': 'yes'},
    {'type': 'end_group'},
]

# Create survey dataframe
survey_df = pd.DataFrame(survey_data)

# Define choices
choices_data = [
    {'list_name': 'gender', 'name': 'male', 'label': 'Male'},
    {'list_name': 'gender', 'name': 'female', 'label': 'Female'},
    {'list_name': 'gender', 'name': 'other', 'label': 'Other'},
    
    {'list_name': 'marital_status', 'name': 'single', 'label': 'Single'},
    {'list_name': 'marital_status', 'name': 'married', 'label': 'Married'},
    {'list_name': 'marital_status', 'name': 'divorced', 'label': 'Divorced'},
    {'list_name': 'marital_status', 'name': 'widowed', 'label': 'Widowed'},
    
    {'list_name': 'education_level', 'name': 'high_school', 'label': 'High School'},
    {'list_name': 'education_level', 'name': 'associates', 'label': 'Associate Degree'},
    {'list_name': 'education_level', 'name': 'bachelors', 'label': 'Bachelor\'s Degree'},
    {'list_name': 'education_level', 'name': 'masters', 'label': 'Master\'s Degree'},
    {'list_name': 'education_level', 'name': 'doctorate', 'label': 'Doctorate'},
]

choices_df = pd.DataFrame(choices_data)

# Define settings
settings_data = [{
    'form_title': 'Employee Details - TIN Auto-fill',
    'form_id': 'employee_details_tin_autofill',
    'version': '2026010401',
    'instance_name': 'concat(${full_name}, " - ", ${tin})'
}]

settings_df = pd.DataFrame(settings_data)

# Create Excel workbook
wb = Workbook()

# Remove default sheet
if 'Sheet' in wb.sheetnames:
    wb.remove(wb['Sheet'])

# Add survey sheet
ws_survey = wb.create_sheet('survey')
for r in dataframe_to_rows(survey_df, index=False, header=True):
    ws_survey.append(r)

# Add choices sheet
ws_choices = wb.create_sheet('choices')
for r in dataframe_to_rows(choices_df, index=False, header=True):
    ws_choices.append(r)

# Add settings sheet
ws_settings = wb.create_sheet('settings')
for r in dataframe_to_rows(settings_df, index=False, header=True):
    ws_settings.append(r)

# Add external_choices sheet - this tells ODK to look for staff_list.csv as attachment
ws_external = wb.create_sheet('external_choices')
ws_external.append(['list_name'])
ws_external.append(['staff_list'])

# Save the workbook
output_file = '../csv/employee_details_odk_form.xlsx'
wb.save(output_file)

print(f"✓ Created {output_file}")
print("\nIMPORTANT: You need to attach 'staff_list.csv' as a media file when uploading to ODK Central")
print("\nThe CSV file should have these columns:")
print("- tin (for lookup)")
print("- eeno (employee number)")
print("- staff_name (full name)")
print("- gender")
print("- organisation")
print("- job")
print("\nNext steps:")
print("1. Rename 'staff-list-with-gender.csv' to 'staff_list.csv'")
print("2. Upload the form to ODK Central")
print("3. Attach 'staff_list.csv' as a media file to the form")
