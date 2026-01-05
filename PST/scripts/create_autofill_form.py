"""
Create ODK XLSForm with external CSV data for auto-populating staff information
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Read the form definition
survey_df = pd.read_csv('employee_odk_form_updated.csv')

# Read choices
choices_df = pd.read_csv('employee_odk_choices.csv')

# Read settings
settings_df = pd.read_csv('employee_odk_settings.csv')

# Update settings to reference external data
settings_df['form_title'] = 'Employee Details with Auto-fill'
settings_df['version'] = '2026010401'

# Create Excel workbook
wb = Workbook()

# Remove default sheet
if 'Sheet' in wb.sheetnames:
    wb.remove(wb['Sheet'])

# Add survey sheet
ws_survey = wb.create_sheet('survey')
for r in dataframe_to_rows(survey_df, index=False, header=True):
    ws_survey.append(r)

# Add choices sheet
ws_choices = wb.create_sheet('choices')
for r in dataframe_to_rows(choices_df, index=False, header=True):
    ws_choices.append(r)

# Add settings sheet
ws_settings = wb.create_sheet('settings')
for r in dataframe_to_rows(settings_df, index=False, header=True):
    ws_settings.append(r)

# Add external_choices sheet for CSV attachment reference
ws_external = wb.create_sheet('external_choices')
ws_external.append(['name'])
ws_external.append(['staff_list'])

# Save the workbook
output_file = 'employee_details_odk_form_autofill.xlsx'
wb.save(output_file)

print(f"✅ Created {output_file}")
print(f"\n📋 Next steps:")
print(f"1. Upload '{output_file}' to ODK Central")
print(f"2. Attach 'staff_list.csv' as a form attachment (media file)")
print(f"3. The form will auto-populate: staff_name, gender, organisation and job position")
print(f"4. Staff only need to enter their TIN (eeno)")
