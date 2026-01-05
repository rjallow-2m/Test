import openpyxl

# Load the workbook
workbook_path = r'c:\Users\Rohey\Test\PST\csv\Employee_Details_ODK_Form.xlsx'
wb = openpyxl.load_workbook(workbook_path)

# Access the settings sheet
if 'settings' in wb.sheetnames:
    ws = wb['settings']
    
    # Find the instance_name column
    instance_name_col = None
    for col in range(1, ws.max_column + 1):
        cell_value = ws.cell(row=1, column=col).value
        if cell_value == 'instance_name':
            instance_name_col = col
            break
    
    if instance_name_col:
        # Delete the column
        ws.delete_cols(instance_name_col)
        print(f"Deleted 'instance_name' column from position {instance_name_col}")
        
        # Save the workbook
        wb.save(workbook_path)
        print(f"Successfully saved {workbook_path}")
    else:
        print("'instance_name' column not found in the settings sheet")
else:
    print("'settings' sheet not found in the workbook")
